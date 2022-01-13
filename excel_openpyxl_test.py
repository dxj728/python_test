# coding=utf-8
# File: excel_openpyxl_test.py
# Author: dxj728
# Time: 2019年03月21日23时
# 说明：openpyxl操作excel的练习
# 参考来源：https://www.cnblogs.com/zeke-python-road/p/8986318.html
# openpyxl支持excel2007以后的版本(.xlsx)
# openpyxl模块安装：pip install openpyxl


import openpyxl
import datetime, time
import sys


# reload(sys)
# sys.setdefaultencoding('utf-8')
# 创建excel操作类
class Excel(object):
    # 初始化表格路径
    def __init__(self, excel_path=None):
        if excel_path:
            self.excel_path = excel_path
        else:
            print('未提供表格路径，可能会出现错误')
            exit(-1)
    
    # 通过提前对文件进行打开测试，查看文件是否可读写
    # 输入：excel文件绝对路径，
    # 返回：无、异常
    def open_excel_check(self, excel_path=None):
        if excel_path is None:
            excel_path = self.excel_path
        try:
            wb = openpyxl.load_workbook(excel_path)
            wb.save(excel_path)
        except:
            print('{} 文件打开测试失败,检查文件是否存在或未关闭'.format(excel_path))
            exit(-1)
        return 0
    
    # 新创建一个excel工作簿并写入表格
    def create_write(self, sheet=None, data=None):
        wb = openpyxl.Workbook()  # 创建文件对象
        if sheet is None:  # 判断是否提供新建表格名
            ws = wb.active  # 创建默认表格（Sheet）
        else:
            ws = wb.create_sheet(sheet)
        # TODO:根据数据类型写入数据（string,list(dict)）
        
        wb.save(self.excel_path)  # 存储内容至excel工作簿
        return
    
    # 加载一个已存在的数据表格并写入数据
    # 输入：sheet(字符串，表格名)
    def load_write(self, sheet=None, data=None):
        wb = openpyxl.load_workbook(self.excel_path)  # 创建文件对象
        if sheet is None:  # 判断是否提供新建表格名
            ws = wb.active  # 获取默认的已存在活动表格
        else:
            ws = wb[sheet]
        # TODO:根据数据类型写入数据（string,list(dict)）
        x = type(data)
        print(x)
        # if type(data) ==
        wb.save(self.excel_path)  # 存储内容至excel工作簿
        return
        pass
    
    # 获取工作簿下该表格全部数据，并返回为list(dict)形式(按照第一行表头为key，从第二行开始读取数据)
    # 输入：sheet(字符串，表格名)
    def load_all_data(self, sheet=None):
        wb = openpyxl.load_workbook(self.excel_path)  # 创建文件对象
        if sheet is None:  # 判断是否提供新建表格名
            ws = wb.active  # 获取默认的已存在活动表格
        else:
            ws = wb[sheet]
        rows = ws.rows
        col = ws.columns
        print(rows, col)
        pass
    
    # 查询并返回表格中的一个单元格中内容
    def load_one_data(self, sheet=None, row=1, col=1):
        wb = openpyxl.load_workbook(self.excel_path)
        wb.guess_types = True
        if sheet is None:
            ws = wb.active
        else:
            ws = wb[sheet]
        x = ws.cell(row, col).value
        wb.save(self.excel_path)
        return x

    def line_chart(self):
        '''折线图学习'''
        from openpyxl.chart import LineChart, Reference
        wb = openpyxl.Workbook()
        ws = wb.active
        rows = [
            ['Date', 'Batch 1', 'Batch 2', 'Batch 3'],
            ['a', 40, 30, 25],
            ['b', 40, 25, 30],
            ['c', 50, 30, 45],
            ['d', 30, 25, 40],
            ['e', 25, 35, 30],
            ['f', 20, 40, 35],
        ]
        for row in rows:
            ws.append(row)
        c1 = LineChart()    # 初始化折线图类
        c1.title = "Line Chart"     # 折线图标题
        c1.style = 13   # 折线图样式
        c1.y_axis.title = 'Size'    # Y轴标题
        c1.x_axis.title = 'Test Number'     # X轴标题
        data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=7)    # 设置数据来源，默认以列区分
        c1.add_data(data, titles_from_data=True)    # 折线图中增加数据
        ws.add_chart(c1, "A10")     # 设置折线图保存位置
        wb.save(r"D:\desktop\line.xlsx")    # 保存后生效

    # 表格的各种奇怪方法尝试练习
    def test(self, data=None):
        wb = openpyxl.load_workbook(self.excel_path)  # 创建文件对象
        ws = wb.active  # 获取默认的第一个sheet
        
        ws['A1'] = 42  # 写入数字
        ws['B1'] = "你好" + "dxj728"  # 写入文字（UTF-8格式）
        ws['A2'] = datetime.datetime.now()  # 写入当前时间
        ws['A3'] = time.strftime("%Y{y}%m{m}%d{d} %H{h}%M{f}%S{s}").format(y='年', m='月', d='日', h='时', f='分',
                                                                           s='秒')  # 写入当前可读时间（字符串）
        ws.cell(3, 3).hyperlink = 'https://www.dxj728.com'  # 插入超链接（明文）
        ws.cell(4, 4, value=('=HYPERLINK("%s"，"%s")' % ('https://www.dxj728.com', 'dxj728')))  # 插入超链接字符
        
        x = ws.cell(6, 6).value
        print(x)
        y = ws.cell(7, 1).value
        z = ws.cell(7, 2).value
        m = ws.cell(6, 1).value
        n = ws.cell(6, 2).value
        
        wb.save(self.excel_path)  # 存储内容至excel工作簿


if __name__ == '__main__':
    excel_path = r'D:\desktop\test.xlsx'
    
    excel_test = Excel(excel_path)
    # excel_test.open_excel_check()
    # excel_test.test()
    # x = excel_test.load_one_data(row=6, col=8)        # unicode类型
    # y = x.split('【')[1].split('】')[0]        # unicode类型
    # z = y.encode('utf-8')        # str类型
    # if y == z:
    #     print('OK')        # 可成功比较
    # print(x)
    # print(y)
    # print(z)
    # excel_test.load_all_data()
    excel_test.test()
