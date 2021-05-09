# --coding:utf-8--
# File: test.py
# Author: dxj728
# Time: 2019年06月03日06时
# 说明：	命令行执行python -m pydoc -p 端口号 ，可以直接查询官方模块文档。

import  sys
import datetime, time
# reload(sys)
# sys.setdefaultencoding('utf-8')


# excel_path = r'D:\desktop\test.xlsx'
#
# ##**********************************读取操作*******************************
# wb = openpyxl.load_workbook(excel_path)
# # ws00 = wb.active
# ws01 = wb['Sheet1']
#
# # 索引从0开始
# x = ws01.cell(4, 2).value
# print(x)
#
# # ********************************写入操作********************************
# ws02 = wb.create_sheet('sheet3_openpyxl')
#
# ws02['A1'] = 42  # 写入数字
# ws02['B1'] = u"你好" + "dxj728"  # 写入文字（UTF-8格式）
# ws02['A2'] = datetime.datetime.now()  # 写入当前时间
# ws02.cell(3, 3).hyperlink = 'https://www.dxj728.com'  # 插入超链接（明文）
# ws02.cell(4, 4, value=('=HYPERLINK("{}","{}")'.format('https://www.dxj728.com', 'dxj728')))  # 插入超链接字符

# wb.save(excel_path)
# 获取所有行
# rows = []
# row1_values = []
# print(ws[0])
# for row in ws.iter_rows():
# 	rows.append(row)
# print(len(rows[0]))
# for i in range(len(rows)):
# 	row1_values.append(rows[0][i].value.encode('utf-8'))
# print(row1_values)
# # 获取所有列
# cols = []
# for col in ws.iter_cols():
# 	cols.append(col)
# print("rows：{} cols:{}".format(rows[0],cols[0]))


# 创建5层嵌套字典dict

# from time_test import dingshi
# if __name__ == '__main__':
# 	end_time = '23:09:10'
# 	ds = dingshi()
# 	ds.start(end_time)
# 	print('ok')

# import os
# from scapy.all import *
#
# time_list = []
# packets = rdpcap(r"D:\desktop\p1.pcap")
# for data in packets:
# 	if data.haslayer("ICMP"):
# 		str_1 = repr(data)
# 		# print(str_1)
# 		# print(ls(data))
# 		time_list.append(data.time)
# 		# print(data['IP'].payload.original)  # 打印出'IP','IPV6','ARP'或者其他
# 		# print(data.payload.name)#打印出'IP','IPV6','ARP'或者其他
#
# # with PcapReader(r"D:\desktop\p1.pcap") as packets: #Windows这样需要考虑转义字符
# # 	x = packets.read_all()
# # 	for data in packets.read_all():
# # 		str_1 = repr(data)
# # 		print(str_1)
#
# print(time_list)
# t1 = 1583844517.933096
#
# x = time_list[1] - time_list[0]
# print(str(x*1000) + 'ms')

# import subprocess
# # Fetches the list of all usb devices:
# result = subprocess.run(['devcon', 'hwids', '=usb'], capture_output=True, text=True)
# print(result)

# # ... add code to parse the result and get the hwid of the device you want ...
#
# subprocess.run(['devcon', 'disable', parsed_hwid]) # to disable
# subprocess.run(['devcon', 'enable', parsed_hwid]) # to enable

# x = [1,2,3,4]
# y = x
# x.pop()
# print(x)
# print(y)

# class grade:
# 	pass
#
#
#
# if __name__ == '__main__':
# 	n, m = input().strip().split()
# 	grade_list = map(int, input().strip().split())
# 	operate_list = [input() for x in range(0, m)]
# 	print(n, m)
# 	print(grade_list)
# 	print(operate_list)

# def deal_str(str1):
# 	ret_list = []
# 	if len(str1) <= 8:
# 		ret_list.append(str1.ljust(8, '0'))
# 		return ret_list
# 	else:
# 		n = int(len(str1) / 8)
# 		for i in range(n):
# 			ret_list.append(str1[i*8:i*8+8])
# 		if len(str1) % 8 != 0:
# 			y = len(str1) % 8
# 			str2 = str1[8*n:]
# 			ret_list.append(str2.ljust(8, '0'))
# 		return ret_list
#
# if __name__ == '__main__':
# 	str1 = input().strip()
# 	if len(str1) == 0:
# 		print(str1)
# 	else:
# 		ret_list = deal_str(str1)
# 		print('\n'.join(map(str, ret_list)))

# import os, sys
# if os.path.exists('1.txt'):
#     print('存在')
#     time.sleep(3)
#     sys.exit(0)
# else:
#     print('无')
#     time.sleep(3)
#     sys.exit(-1)

# cb_en = 'python编程语言'.encode('utf-8')
# print(cb_en)        # b'python\xe7\xbc\x96\xe7\xa8\x8b\xe8\xaf\xad\xe8\xa8\x80'     \\ 每个数据单元为1个字节(8位),\x代表16进制,e和7代表2个十六进制数(各4位)
# cb_de = cb_en.decode(encoding='utf-8')
# print(cb_de)        # python编程语言
#
# en=b'\x90'


# def get_hight(height, count):
#     sum = 2 ** height
#     while sum < count:
#         height = height + 1
#         temp = 2 ** height
#         sum = sum + temp
#     return height + 1
#
# def get_result(test_list):
#     height = 0
#     start = test_list[0:1]
#     result_list = []
#     result_list.append(start)
#     sum = 1
#     while sum < len(test_list):
#         height = height + 1
#         temp = 2 ** height
#         start = sum
#         end = sum + temp
#         if end > len(test_list):
#             end = len(test_list) - 1
#         temp_list = test_list[start:end]
#         for i in range(0, len(temp_list)):
#             if None in temp_list:
#                 temp_list.remove(None)
#         result_list.append(temp_list)
#         sum = sum + temp
#     return result_list
#
# if __name__ == '__main__':
#     test_list = [3, 9, 20, None, None, 15, 7]
#     # print(len(test_list))
#     # height = get_hight(0, 7)
#     # print(height)
#     # start_node = test_list[0]
#     result_list = get_result(test_list)
#     print(result_list)


# import socket
#
# s = socket.socket()
# s.connect(('127.0.0.1', 30000))
# print('---{}---'.format(s.recv(1024).decode('utf-8')))
# s.close()

import openpyxl
from openpyxl.drawing.image import Image

def insert_image(insert_location, image_path):

    # # 要插入的图片路径、插入后的图片尺寸
    # print(image_path)

    img = Image(image_path)
    new_size = (150, 80)
    img.width, img.height = new_size

    sheet[insert_location] = ""
    sheet.add_image(img, insert_location)

    wb.save(xlsx_path)
    print("插入成功!")

xlsx_path = r'D:\desktop\teste.xlsx'
image_absolute_path = r'D:\desktop\1.jpg'

wb = openpyxl.load_workbook(xlsx_path)
sheet = wb["Sheet1"]

# FileObj = xlrd.open_workbook(xlsx_path)
# # 查找id_number所在的位置
# sheet_x = FileObj.sheet_by_name("基本信息表")
# row_count = sheet_x.nrows
# lists = sheet_x.col_values(10)


insert_image("D3", image_absolute_path)
















